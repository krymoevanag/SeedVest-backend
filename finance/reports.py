"""Binary financial-report generators used by the finance reporting endpoints."""

from decimal import Decimal
from io import BytesIO
from xml.sax.saxutils import escape

from django.db.models import Sum

from .models import Contribution, Loan, MonthlyContributionRecord, Penalty


def _decimal(value):
    return Decimal(value or "0.00")


def _member_name(user):
    return f"{user.first_name} {user.last_name}".strip() or user.email


def build_financial_cycle_workbook(cycle):
    """Return an in-memory XLSX ledger and balance-sheet workbook for a cycle."""
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Font, PatternFill
    from openpyxl.utils import get_column_letter

    contributions = (
        Contribution.objects.filter(financial_cycle=cycle, is_archived=False)
        .select_related("user", "group")
        .order_by("due_date", "user__first_name", "user__last_name")
    )
    monthly_records = (
        MonthlyContributionRecord.objects.filter(financial_cycle=cycle, is_archived=False)
        .select_related("user")
        .order_by("month", "user__first_name", "user__last_name")
    )
    loans = (
        Loan.objects.filter(financial_cycle=cycle, is_archived=False)
        .select_related("user")
        .order_by("created_at")
    )
    penalties = Penalty.objects.filter(
        contribution__financial_cycle=cycle,
        is_archived=False,
    ).select_related("user", "contribution")

    paid_total = _decimal(
        contributions.filter(status__in=["PAID", "LATE"]).aggregate(total=Sum("amount"))["total"]
    )
    expected_total = _decimal(monthly_records.aggregate(total=Sum("expected_contribution_amount"))["total"])
    outstanding_total = _decimal(monthly_records.aggregate(total=Sum("outstanding_amount"))["total"])
    penalty_total = _decimal(penalties.aggregate(total=Sum("amount"))["total"])
    active_loan_balance = _decimal(
        loans.filter(status="DISBURSED").aggregate(total=Sum("balance_remaining"))["total"]
    )

    workbook = Workbook()
    summary = workbook.active
    summary.title = "Cycle Summary"
    header_fill = PatternFill("solid", fgColor="1B5E20")
    header_font = Font(color="FFFFFF", bold=True)

    summary.append(["SeedVest financial cycle report"])
    summary.append(["Group", cycle.group.name])
    summary.append(["Cycle", cycle.cycle_name])
    summary.append(["Period", f"{cycle.start_date:%d %b %Y} - {cycle.end_date:%d %b %Y}"])
    summary.append(["Status", cycle.status])
    summary.append([])
    summary.append(["Metric", "KES"])
    for cell in summary[7]:
        cell.fill = header_fill
        cell.font = header_font
    summary.append(["Contributions collected", paid_total])
    summary.append(["Contributions expected", expected_total])
    summary.append(["Outstanding contributions", outstanding_total])
    summary.append(["Penalties issued", penalty_total])
    summary.append(["Outstanding disbursed loans", active_loan_balance])
    summary.append(["Net savings before investments", paid_total - penalty_total])
    summary.column_dimensions["A"].width = 34
    summary.column_dimensions["B"].width = 24

    ledger = workbook.create_sheet("Contribution Ledger")
    ledger_headers = [
        "Member",
        "Email",
        "Due date",
        "Paid date",
        "Expected amount",
        "Amount received",
        "Status",
        "Reference",
    ]
    ledger.append(ledger_headers)
    for cell in ledger[1]:
        cell.fill = header_fill
        cell.font = header_font
    for contribution in contributions:
        ledger.append(
            [
                _member_name(contribution.user),
                contribution.user.email,
                contribution.due_date,
                contribution.paid_date,
                contribution.expected_amount,
                contribution.amount,
                contribution.status,
                contribution.reported_reference,
            ]
        )

    loans_sheet = workbook.create_sheet("Loan Ledger")
    loan_headers = [
        "Borrower",
        "Principal",
        "Interest rate (%)",
        "Duration (months)",
        "Total payable",
        "Balance remaining",
        "Status",
        "Due date",
    ]
    loans_sheet.append(loan_headers)
    for cell in loans_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    for loan in loans:
        loans_sheet.append(
            [
                _member_name(loan.user),
                loan.amount,
                loan.interest_rate,
                loan.duration_months,
                loan.total_payable,
                loan.balance_remaining,
                loan.status,
                loan.due_date,
            ]
        )

    balance_sheet = workbook.create_sheet("Balance Sheet")
    balance_sheet.append(["Balance Sheet", "KES"])
    for cell in balance_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
    balance_sheet.append(["Assets: cash contributions collected", paid_total])
    balance_sheet.append(["Assets: loans receivable", active_loan_balance])
    balance_sheet.append(["Liabilities: outstanding member contributions", outstanding_total])
    balance_sheet.append(["Net position", paid_total + active_loan_balance - outstanding_total])

    for sheet in workbook.worksheets:
        sheet.freeze_panes = "A2" if sheet.max_row > 1 else None
        for column_index in range(1, sheet.max_column + 1):
            width = min(
                42,
                max(
                    14,
                    max(
                        len(str(cell.value)) if cell.value is not None else 0
                        for cell in sheet[get_column_letter(column_index)]
                    )
                    + 2,
                ),
            )
            sheet.column_dimensions[get_column_letter(column_index)].width = width
        for row in sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top")

    stream = BytesIO()
    workbook.save(stream)
    stream.seek(0)
    return stream


def build_member_statement_pdf(member, group, cycle=None):
    """Return a formatted PDF statement for one member in the selected group."""
    from reportlab.lib import colors
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib.units import mm
    from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

    contributions = Contribution.objects.filter(
        user=member,
        group=group,
        is_archived=False,
    ).order_by("due_date", "created_at")
    loans = Loan.objects.filter(user=member, group=group, is_archived=False).order_by("created_at")
    penalties = Penalty.objects.filter(user=member, is_archived=False).filter(
        contribution__group=group
    )
    if cycle is not None:
        contributions = contributions.filter(financial_cycle=cycle)
        loans = loans.filter(financial_cycle=cycle)
        penalties = penalties.filter(contribution__financial_cycle=cycle)

    saved_total = _decimal(
        contributions.filter(status__in=["PAID", "LATE"]).aggregate(total=Sum("amount"))["total"]
    )
    outstanding_total = _decimal(
        contributions.exclude(status__in=["PAID", "LATE"]).aggregate(total=Sum("expected_amount"))["total"]
    )
    penalties_total = _decimal(penalties.aggregate(total=Sum("amount"))["total"])
    loan_balance = _decimal(loans.exclude(status__in=["REPAID", "REJECTED"]).aggregate(total=Sum("balance_remaining"))["total"])

    stream = BytesIO()
    document = SimpleDocTemplate(
        stream,
        pagesize=A4,
        rightMargin=16 * mm,
        leftMargin=16 * mm,
        topMargin=16 * mm,
        bottomMargin=16 * mm,
    )
    styles = getSampleStyleSheet()
    story = [
        Paragraph("SeedVest member financial statement", styles["Title"]),
        Spacer(1, 6 * mm),
        Paragraph(f"<b>Member:</b> {escape(_member_name(member))}", styles["BodyText"]),
        Paragraph(f"<b>Group:</b> {escape(group.name)}", styles["BodyText"]),
        Paragraph(
            f"<b>Cycle:</b> {escape(cycle.cycle_name) if cycle else 'All available records'}",
            styles["BodyText"],
        ),
        Spacer(1, 6 * mm),
    ]
    summary_data = [
        ["Summary", "KES"],
        ["Contributions received", f"{saved_total:,.2f}"],
        ["Outstanding contributions", f"{outstanding_total:,.2f}"],
        ["Penalties", f"{penalties_total:,.2f}"],
        ["Outstanding loan balance", f"{loan_balance:,.2f}"],
    ]
    summary_table = Table(summary_data, colWidths=[112 * mm, 55 * mm])
    summary_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#1B5E20")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#B0BEC5")),
                ("ALIGN", (1, 1), (1, -1), "RIGHT"),
                ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F7FA")]),
            ]
        )
    )
    story.extend([summary_table, Spacer(1, 8 * mm)])

    story.append(Paragraph("Contribution history", styles["Heading2"]))
    contribution_rows = [["Due", "Paid", "Expected", "Received", "Status"]]
    for contribution in contributions:
        contribution_rows.append(
            [
                contribution.due_date.strftime("%d %b %Y"),
                contribution.paid_date.strftime("%d %b %Y") if contribution.paid_date else "-",
                f"{contribution.expected_amount:,.2f}",
                f"{contribution.amount:,.2f}",
                contribution.status,
            ]
        )
    if len(contribution_rows) == 1:
        contribution_rows.append(["No contribution records", "", "", "", ""])
    contribution_table = Table(contribution_rows, repeatRows=1)
    contribution_table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#102A43")),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("GRID", (0, 0), (-1, -1), 0.25, colors.HexColor("#CFD8DC")),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
                ("ALIGN", (2, 1), (3, -1), "RIGHT"),
            ]
        )
    )
    story.append(contribution_table)
    document.build(story)
    stream.seek(0)
    return stream